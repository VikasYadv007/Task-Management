from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import io

from .models import Task
from .serializers import TaskSerializer

class TaskViewSet(viewsets.ModelViewSet):
    """
    API endpoint for tasks with CRUD operations and Excel export
    """
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        This view returns a list of all tasks for the authenticated user.
        """
        user_id = self.request.query_params.get('user_id', None)
        if user_id:
            return Task.objects.filter(user_id=user_id).order_by('-created_at')
        return Task.objects.all()

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export tasks to Excel
        """
        user_id = self.request.query_params.get('user_id', None)
        
        if not user_id:
            return Response({'error': 'User ID is required'}, status=status.HTTP_400_BAD_REQUEST)
            
        tasks = Task.objects.filter(user_id=user_id).order_by('-created_at')
        
        # Create a workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Tasks"
        
        # Add headers
        headers = ['ID', 'Title', 'Description', 'Effort (Days)', 'Due Date', 'Created At']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add data
        for row_num, task in enumerate(tasks, 2):
            worksheet.cell(row=row_num, column=1).value = task.id
            worksheet.cell(row=row_num, column=2).value = task.title
            worksheet.cell(row=row_num, column=3).value = task.description
            worksheet.cell(row=row_num, column=4).value = task.effort_days
            worksheet.cell(row=row_num, column=5).value = task.due_date.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=6).value = task.created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # Create a response with Excel file
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=tasks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
